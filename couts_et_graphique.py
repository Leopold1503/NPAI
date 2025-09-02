import os
import unicodedata
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# ================== PARAMÈTRES ==================
FICHIER_2024 = r"U:\Business Assurance\Revenue Assurance\RA_2024\FE2026\PND-NPAI\Fichiers Asterion\B2C\Etude Asterion 2025\SFR-concaténation20250204.xlsx"
FEUILLE_2024 = "20240101-20241216"   # feuille 2024 à lire

FICHIER_2025 = r"U:\Business Assurance\Revenue Assurance\RA_2024\FE2026\PND-NPAI\Fichiers Asterion\B2C\Etude Asterion 2025\NPAI 2025.xlsx"
FEUILLE_2025 = None                  # None => auto-détection de la feuille contenant les colonnes attendues

FICHIER_SORTIE = r"U:\Business Assurance\Revenue Assurance\RA_2024\FE2026\PND-NPAI\Fichiers Asterion\B2C\Etude Asterion 2025\Frais documents.xlsx"
IMAGE_GRAPHE  = r"U:\Business Assurance\Revenue Assurance\RA_2024\FE2026\PND-NPAI\Fichiers Asterion\B2C\Etude Asterion 2025\Evolution_traitements.png"

# Tarifs par type
TARIFS = {
    "FACTURE":   0.75,
    "RELANCE":   0.75,
    "COURRIER":  0.81,
    "DUPLICATA": 0.75,  # même que facture
}

# Ordre d’affichage des lignes
ORDRE_TYPES = ["FACTURE", "RELANCE", "COURRIER", "DUPLICATA"]

# Mois FR pour colonnes et axes
MOIS_FR = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
           "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]


# ================== OUTILS ==================
def _strip_accents_lower(s: str) -> str:
    """minuscule + sans accents + espaces normalisés (pour matcher les noms de colonnes)."""
    if not isinstance(s, str):
        s = str(s)
    s = unicodedata.normalize("NFD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    s = " ".join(s.split())
    return s.lower().strip()

def normaliser_type(type_brut: str) -> str:
    """
    Mappe les libellés vers FACTURE / RELANCE / COURRIER / DUPLICATA.
    Tolérant aux accents/variantes (Relançe, Dupli…, Courrier simple, Facture PDF, etc.)
    """
    t = unicodedata.normalize("NFD", str(type_brut)).encode("ascii","ignore").decode("ascii")
    t = t.upper().strip()
    if "DUPLI" in t:
        return "DUPLICATA"
    if "FACT" in t:
        return "FACTURE"
    if "RELAN" in t:
        return "RELANCE"
    if "COURR" in t:
        return "COURRIER"
    return t  # inconnu -> ignoré si non tarifé


def trouver_colonnes(df: pd.DataFrame, nom_type="TYPE DE DOCUMENT", nom_date="DATE TRAITEMENT PND"):
    """
    Retrouve les colonnes par NOMS (tolérant : casse/accents/espaces).
    Renvoie (col_type, col_date) ou lève une erreur claire.
    """
    colmap = {_strip_accents_lower(c): c for c in df.columns.astype(str)}
    cible_type = _strip_accents_lower(nom_type)
    cible_date = _strip_accents_lower(nom_date)

    col_type = colmap.get(cible_type)
    col_date = colmap.get(cible_date)

    def cherche_approx(cible, tokens_min):
        for cnorm, corig in colmap.items():
            if all(tok in cnorm for tok in tokens_min):
                return corig
        return None

    if col_type is None:
        col_type = cherche_approx(cible_type, ["type", "document"])
    if col_date is None:
        col_date = (cherche_approx(cible_date, ["date", "traitement", "pnd"])
                    or cherche_approx(cible_date, ["date", "pnd"]))

    if col_type is None or col_date is None:
        raise ValueError(
            "Colonnes introuvables. Requis : "
            f"'{nom_type}' et '{nom_date}'. Colonnes disponibles : {list(df.columns)}"
        )
    return col_type, col_date

def lire_feuille(chemin_xlsx: str, sheet_name: str | None) -> pd.DataFrame:
    """
    Lit une feuille précise si sheet_name est fourni.
    Sinon, essaie toutes les feuilles et renvoie la première qui contient les colonnes attendues.
    """
    if not os.path.exists(chemin_xlsx):
        raise FileNotFoundError(f"Fichier introuvable : {chemin_xlsx}")

    if sheet_name:
        df = pd.read_excel(chemin_xlsx, sheet_name=sheet_name, engine="openpyxl")
        if df.empty:
            raise ValueError(f"La feuille '{sheet_name}' est vide dans : {chemin_xlsx}")
        return df

    # Auto-détection
    xls = pd.ExcelFile(chemin_xlsx, engine="openpyxl")
    last_err = None
    for sh in xls.sheet_names:
        try:
            dfi = pd.read_excel(chemin_xlsx, sheet_name=sh, engine="openpyxl")
            if dfi.empty:
                continue
            # test : trouver les colonnes attendues (ne lève pas si ok)
            trouver_colonnes(dfi)
            return dfi
        except Exception as e:
            last_err = e
            continue
    raise ValueError(
        f"Aucune feuille ne contient les colonnes attendues dans : {chemin_xlsx}. "
        f"Dernière erreur : {last_err}"
    )

def lire_fichier(chemin_xlsx: str, sheet_name: str | None) -> pd.DataFrame:
    """
    Lit l'Excel, récupère 'TYPE DE DOCUMENT' et 'DATE TRAITEMENT PND' par NOM,
    nettoie/normalise, puis ajoute Year/Mois. Renvoie un DF standardisé.
    """
    df_raw = lire_feuille(chemin_xlsx, sheet_name)
    col_type, col_date = trouver_colonnes(df_raw, "TYPE DE DOCUMENT", "DATE TRAITEMENT PND")

    out = pd.DataFrame({
        "TypeDocument": df_raw[col_type].astype(str).str.strip(),
        "Date": pd.to_datetime(df_raw[col_date], errors="coerce", dayfirst=True)
    }).dropna(subset=["TypeDocument", "Date"])

    out["TypeNorm"] = out["TypeDocument"].apply(normaliser_type)
    out["Year"]  = out["Date"].dt.year.astype(int)
    out["Month"] = out["Date"].dt.month.astype(int)
    return out


def frais_par_annee(df_total: pd.DataFrame, annee: int) -> pd.DataFrame:
    """
    Tableau (lignes=type, colonnes=mois Jan-Déc, + 'Total annuel') pour une année donnée.
    Calcul = nombre de lignes * tarif, par type et par mois.
    + Ajoute une ligne 'TOTAL (3 types)' = FACTURE + RELANCE + COURRIER
    """
    sous = df_total[df_total["Year"] == annee].copy()
    sous = sous[sous["TypeNorm"].isin(TARIFS.keys())]

    counts = sous.groupby(["TypeNorm", "Month"]).size().unstack(fill_value=0)
    # assure 12 mois
    for m in range(1, 13):
        if m not in counts.columns:
            counts[m] = 0
    counts = counts[sorted(counts.columns)]  # 1..12

    # coûts par type
    costs = counts.mul(pd.Series(TARIFS), axis=0).round(2)

    # renomme colonnes en mois FR et ajoute total annuel
    costs.columns = MOIS_FR
    costs["Total annuel"] = costs.sum(axis=1).round(2)

    # ordonner les lignes
    ordre = [t for t in ORDRE_TYPES if t in costs.index]
    costs = costs.reindex(ordre)

    # === Ajouter la ligne de totaux mensuels sur 3 types (FACTURE + RELANCE + COURRIER) ===
    types_3 = [t for t in ["FACTURE", "RELANCE", "COURRIER"] if t in costs.index]
    if types_3:
        total_row = costs.loc[types_3].sum(axis=0)
        costs.loc["TOTAL (3 types)"] = total_row

    return costs


def evolution_traitements(df_total: pd.DataFrame, annees=(2024, 2025)) -> pd.DataFrame:
    """
    Tableau (index=année, colonnes=mois FR) avec le nombre de traitements (toutes lignes).
    """
    pivot = df_total.groupby(["Year", "Month"]).size().unstack(fill_value=0)
    for m in range(1, 13):
        if m not in pivot.columns:
            pivot[m] = 0
    pivot = pivot[sorted(pivot.columns)]
    pivot = pivot.reindex(index=annees, fill_value=0)
    pivot.columns = MOIS_FR
    return pivot


def tracer_graphe(df_evol: pd.DataFrame, path_png: str = None):
    """Trace les 2 courbes (2024 vs 2025) et enregistre une image si demandé."""
    plt.figure(figsize=(11, 6))
    for an in df_evol.index:
        y = df_evol.loc[an, :].values
        plt.plot(MOIS_FR, y, marker="o", label=str(an))
    plt.title("Évolution mensuelle du nombre de traitements (2024 vs 2025)")
    plt.xlabel("Mois")
    plt.ylabel("Nombre de traitements")
    plt.grid(True)
    plt.xticks(rotation=45)
    plt.legend()
    plt.tight_layout()
    if path_png:
        plt.savefig(path_png, dpi=150, bbox_inches="tight")
    plt.show()


def formater_monnaie_excel(fichier_xlsx: str, feuilles=("Frais 2024", "Frais 2025")):
    """
    Ouvre le fichier Excel et applique un format monétaire (€) aux cellules
    des feuilles de frais (toutes les colonnes de mois + 'Total annuel').
    Hypothèse : l’index (types) est en colonne A, donc montants à partir de colonne B.
    """
    wb = load_workbook(fichier_xlsx)
    fmt = '€ #,##0.00'  # format monétaire simple

    for sheet_name in feuilles:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if ws.max_row < 2 or ws.max_column < 3:
            continue

        # Repérer colonnes numériques : de B à la dernière (mois + Total annuel)
        first_row = 1  # en-têtes
        first_data_row = 2
        first_amount_col = 2  # B
        last_col = ws.max_column
        last_row = ws.max_row

        # Appliquer format sur toutes les cellules numériques (B2:LastCol, LastRow)
        for r in range(first_data_row, last_row + 1):
            for c in range(first_amount_col, last_col + 1):
                cell = ws.cell(row=r, column=c)
                # on force format monétaire (même si la valeur est 0 ou vide)
                cell.number_format = fmt

    wb.save(fichier_xlsx)
    wb.close()


# ================== PIPELINE ==================
def main():
    # 1) Charger 2024 (feuille explicite) et 2025 (auto-détection par colonnes)
    df24 = lire_fichier(FICHIER_2024, FEUILLE_2024)
    df25 = lire_fichier(FICHIER_2025, FEUILLE_2025)

    # 2) Concaténer (la date fait foi => décembre 2024 dans le fichier 2025 est classé en 2024)
    df_total = pd.concat([df24, df25], ignore_index=True)

    # 3) Frais documentaires (tableaux par année, avec ligne TOTAL (3 types))
    frais_2024 = frais_par_annee(df_total, 2024)
    frais_2025 = frais_par_annee(df_total, 2025)

    # 4) Évolution des traitements (volume par mois)
    df_evol = evolution_traitements(df_total, annees=(2024, 2025))

    # 5) Écrire le fichier Excel (3 onglets)
    with pd.ExcelWriter(FICHIER_SORTIE, engine="openpyxl") as writer:
        frais_2024.to_excel(writer, sheet_name="Frais 2024")
        frais_2025.to_excel(writer, sheet_name="Frais 2025")
        df_evol.to_excel(writer, sheet_name="Évolution traitements")

    # 6) Formater en monétaire (€) les feuilles de frais
    formater_monnaie_excel(FICHIER_SORTIE, feuilles=("Frais 2024", "Frais 2025"))

    # 7) Graphe comparatif (deux courbes) + image PNG
    tracer_graphe(df_evol, path_png=IMAGE_GRAPHE)

    print(f"✅ Fichier écrit : {FICHIER_SORTIE}")
    print(f"✅ Graphe enregistré : {IMAGE_GRAPHE}")


if __name__ == "__main__":
    main()
